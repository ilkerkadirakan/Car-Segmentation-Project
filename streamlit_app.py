#
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from pandas.io.xml import preprocess_data
from sklearn.preprocessing import LabelEncoder, MinMaxScaler
from sklearn.cluster import KMeans, AgglomerativeClustering
from sklearn.metrics import silhouette_score, calinski_harabasz_score, davies_bouldin_score
from sklearn.decomposition import PCA
from sklearn.model_selection import ParameterGrid
from scipy.spatial.distance import cdist
import xlsxwriter
from openpyxl import load_workbook
from io import BytesIO
import warnings
warnings.filterwarnings("ignore")

# Başlık
st.title("🚗 Vehicle Customer Segmentation")
st.markdown("""
### **Data Mining Final Project**
#### **Vehicle Customer Segmentation Analysis**
This application showcases the customer segmentation process for vehicles based on their specifications and pricing.

📌 **Project Goals:**
- Identify automobiles that serve similar customer needs.
- Understand consumer preferences based on specifications and pricing.
- Optimize product management and strategic planning for manufacturers.
- Provide data-driven insights for automakers to retain and grow market share.

📖 **Dataset Description:**
The dataset consists of **157 rows** and **26 columns**, containing both numerical and categorical data. It includes vehicle specifications such as price, engine size, horsepower, dimensions, fuel capacity, and efficiency.
""")

# Data Overview
st.write("### Data Overview")
data = pd.read_excel("2.xlsx")
st.dataframe(data.head())

st.write("### Data Dictionary")
st.write("""
| **Column Name** | **Data Type** | **Description** |
|------------------|---------------|------------------|
| `manufact`       | String        | Car manufacturer or brand (e.g., Ford, Honda, BMW). |
| `model`          | String        | Specific model name of the vehicle (e.g., Accord, Explorer). |
| `sales`          | Float         | Number of units sold (in thousands). Represents market demand. |
| `resale`         | Float         | Estimated resale value of the vehicle after a certain period (likely 3-5 years). Indicates value retention. |
| `type`           | Integer       | Car category indicator: 0 = Passenger car (sedan, hatchback, coupe), 1 = Utility/commercial (SUVs, pickups, vans). |
| `price`          | Float         | Manufacturer's suggested retail price (in thousands of dollars). |
| `engine_s`       | Float         | Engine size in liters (e.g., 2.0L, 3.5L). Larger sizes usually indicate more power. |
| `horsepow`       | Float         | Horsepower (HP) of the vehicle. A measure of engine power. |
| `wheelbas`       | Float         | Wheelbase length in inches — the distance between front and rear axles. Affects ride comfort and space. |
| `width`          | Float         | Width of the car in inches. |
| `length`         | Float         | Length of the car in inches. |
| `curb_wgt`       | Float         | Curb weight in pounds — the total weight of the vehicle without passengers or cargo. |
| `fuel_cap`       | Float         | Fuel tank capacity in gallons. |
| `mpg`            | Float         | Fuel efficiency: miles per gallon. Higher is better for economy. |
| `lnsales`        | Float         | Natural log of the sales column. Used for normalization or regression analysis. |
| `zresale`        | Float         | Standardized (z-score) resale value. Shows how many standard deviations each resale value is from the mean. |
| `ztype`          | Float         | Standardized version of type. Not usually meaningful unless used in statistical models. |
| `zprice`         | Float         | Standardized price value. |
| `zengine_`       | Float         | Standardized engine size. |
| `zhorsepo`       | Float         | Standardized horsepower. |
| `zwheelba`       | Float         | Standardized wheelbase. |
| `zwidth`         | Float         | Standardized width. |
| `zlength`        | Float         | Standardized length. |
| `zcurb_wg`       | Float         | Standardized curb weight. |
| `zfuel_ca`       | Float         | Standardized fuel capacity. |
| `zmpg`           | Float         | Standardized fuel efficiency (mpg). |
""")

st.image("graphs/step2_initial_data_distribution.png", use_container_width=True)

numerical_columns = data.select_dtypes(include=['int64', 'float64']).columns.tolist()
categorical_columns = data.select_dtypes(include=['object']).columns.tolist()

st.write(f"Numerical Columns: {len(numerical_columns)}")
st.write(f"Categorical Columns: {len(categorical_columns)}")

st.image("graphs/step2_column_distribution.png", use_container_width=True)




# Eksik Veri Doldurma
st.write("### Handling Missing Data")
missing_summary = data.isnull().sum()
st.write("Missing values summary:")
st.write(missing_summary[missing_summary > 0])


for col in numerical_columns:
    data[col].fillna(data[col].median(), inplace=True)
for col in categorical_columns:
    data[col].fillna(data[col].mode()[0], inplace=True)

st.write("Missing values filled using median (numerical) and mode (categorical). ✅")

from openpyxl import load_workbook

wb = load_workbook("sheets/step3_filled_data_highlighted.xlsx")
ws = wb.active

filled_data = pd.read_excel("sheets/step3_filled_data_highlighted.xlsx", header=0, engine='openpyxl')
background_color_mask = []
for row in ws.iter_rows(min_row=2):  # Başlık satırını atla
    row_colors = []
    for cell in row:
        if cell.fill.start_color.index == 'FF00FF00':  # Yeşil renk kontrolü
            row_colors.append(True)
        else:
            row_colors.append(False)
    background_color_mask.append(any(row_colors))

highlighted_rows = filled_data[background_color_mask]
st.dataframe(highlighted_rows)
st.write("Missing values filled rows are shown in the table above.")

# Kategorik Değişkenleri Kodlama
st.write("### Encoding Categorical Variables")
if 'manufact' in categorical_columns and 'model' in categorical_columns:
    data['manufact_model'] = data['manufact'] + '_' + data['model']
    label_encoder = LabelEncoder()
    data['manufact_model_encoded'] = label_encoder.fit_transform(data['manufact_model'])
    st.write("Combined and encoded manufacturer & model information. ✅")

    st.write("### Encoded Data")
    st.dataframe(data[['manufact', 'model', 'manufact_model', 'manufact_model_encoded']].head())

# Aykırı Değer Temizleme (IQR)
st.write("### Handling Outliers using IQR")
Q1 = data[numerical_columns].quantile(0.25)
Q3 = data[numerical_columns].quantile(0.75)
IQR = Q3 - Q1
for col in numerical_columns:
    lower_bound = Q1[col] - 1.5 * IQR[col]
    upper_bound = Q3[col] + 1.5 * IQR[col]
    data[col] = data[col].clip(lower=lower_bound, upper=upper_bound)

st.write("Distribution of numerical columns before and after outlier handling:")
st.write("Before:")
st.image("graphs/step3_before_cleaning_data_distribution.png", use_container_width=True)
st.write("After:")
st.image("graphs/step3_cleaned_data_distribution.png", use_container_width=True)

st.write("Outliers removed using IQR method. ✅")

st.write("### Data After Preprocessing")
path = "sheets/step3_prepared_data_highlighted.xlsx"
prepared_data = pd.read_excel(path, header=0, engine='openpyxl')
wb=load_workbook(path)
ws=wb.active

background_color_mask = []
for row in ws.iter_rows(min_row=2):  # Skip header row
    row_colors = []
    for cell in row:
        if cell.fill.start_color.index == 'FFFFEB9C':  # Yellow color check
            row_colors.append(True)
        else:
            row_colors.append(False)
    background_color_mask.append(any(row_colors))
highlighted_rows=prepared_data[background_color_mask]
st.dataframe(highlighted_rows)
st.write("Outliers handled rows are shown in the table above.")

preprocess_data=data.copy()
csv_file=preprocess_data.to_csv(index=False)

st.download_button("Download Preprocessed Data",csv_file, "prepared_data.csv")



# Normalize Data
st.write("### Standardizing Data")
scaler = MinMaxScaler()
X_scaled = scaler.fit_transform(data[numerical_columns])
st.write("Data standardized using MinMaxScaler. ✅")

st.write("### Standardized Data")
st.dataframe(pd.DataFrame(X_scaled, columns=numerical_columns).head())

# PCA
st.write("### Principal Component Analysis (PCA)")
pca = PCA(n_components=2)
X_pca = pca.fit_transform(X_scaled)
data['PCA1'] = X_pca[:, 0]
data['PCA2'] = X_pca[:, 1]
st.write("PCA applied to reduce dimensions. ✅")

plt.figure(figsize=(12, 6))
plt.scatter(X_pca[:, 0], X_pca[:, 1], alpha=0.7)
plt.title("PCA Scatter Plot")
st.pyplot(plt)

# K-Means Kümeleme
st.write("### K-Means Clustering")
param_grid_kmeans = {
    'n_clusters': [2, 3, 4, 5, 6, 7, 8, 9, 10],
    'init': ['k-means++', 'random'],
    'n_init': [10, 20],
    'max_iter': [300, 500],
    'algorithm': ['lloyd', 'elkan']
}
best_score_kmeans = -1
best_params_kmeans = None

for params in ParameterGrid(param_grid_kmeans):
    kmeans_temp = KMeans(
        n_clusters=params['n_clusters'],
        init=params['init'],
        n_init=params['n_init'],
        max_iter=params['max_iter'],
        algorithm=params['algorithm'],
        random_state=42
    )
    labels_temp = kmeans_temp.fit_predict(X_pca)
    sil_temp = silhouette_score(X_pca, labels_temp)
    if sil_temp > best_score_kmeans:
        best_score_kmeans = sil_temp
        best_params_kmeans = params

st.write(f"Best Silhouette Score (K-Means) from tuning: {best_score_kmeans:.4f}")
st.write("Best Parameters (K-Means):", best_params_kmeans)

kmeans = KMeans(
    n_clusters=best_params_kmeans['n_clusters'],
    init=best_params_kmeans['init'],
    n_init=best_params_kmeans['n_init'],
    max_iter=best_params_kmeans['max_iter'],
    algorithm=best_params_kmeans['algorithm'],
    random_state=42
)
data['Cluster_KMeans'] = kmeans.fit_predict(X_pca)
silhouette_kmeans = silhouette_score(X_pca, data['Cluster_KMeans'])
st.write(f"**Silhouette Score (K-Means):** {silhouette_kmeans:.4f}")

plt.figure(figsize=(12, 6))
plt.scatter(X_pca[:, 0], X_pca[:, 1], c=data['Cluster_KMeans'], cmap='viridis', alpha=0.7)
plt.title("K-Means Clustering with PCA")
st.pyplot(plt)

# Agglomerative Kümeleme
st.write("### Agglomerative Clustering")
param_grid_agg = {
    'n_clusters': [2, 3, 4, 5, 6, 7, 8, 9, 10],
    'metric': ['euclidean', 'manhattan', 'cosine'],
    'linkage': ['ward', 'complete', 'average', 'single']
}
best_score_agg = -1
best_params_agg = None

for params in ParameterGrid(param_grid_agg):
    if params['linkage'] == 'ward' and params['metric'] != 'euclidean':
        continue

    agg_temp = AgglomerativeClustering(
        n_clusters=params['n_clusters'],
        metric=params['metric'],
        linkage=params['linkage']
    )
    labels_temp = agg_temp.fit_predict(X_pca)
    sil_temp = silhouette_score(X_pca, labels_temp)

    if sil_temp > best_score_agg:
        best_score_agg = sil_temp
        best_params_agg = params

st.write(f"Best Silhouette Score (Agglomerative) from tuning: {best_score_agg:.4f}")
st.write("Best Parameters (Agglomerative):", best_params_agg)

agg = AgglomerativeClustering(
    n_clusters=best_params_agg['n_clusters'],
    metric=best_params_agg['metric'],
    linkage=best_params_agg['linkage']
)
data['Cluster_Agglomerative'] = agg.fit_predict(X_pca)
silhouette_agg = silhouette_score(X_pca, data['Cluster_Agglomerative'])
st.write(f"**Silhouette Score (Agglomerative):** {silhouette_agg:.4f}")

plt.figure(figsize=(12, 6))
plt.scatter(X_pca[:, 0], X_pca[:, 1], c=data['Cluster_Agglomerative'], cmap='plasma', alpha=0.7)
plt.title("Agglomerative Clustering with PCA")
st.pyplot(plt)

# Sonuç Değerlendirme
st.write("### Evaluation and Insights")
st.markdown("""
- **K-Means clustering performed slightly better**, with higher separation among clusters.
- **Agglomerative clustering** showed slightly lower scores but still provided meaningful groupings.
- **Silhouette scores confirm the effectiveness** of both clustering methods.
- **Market segments were identified**, including economy, performance, and luxury vehicle categories.
- **Strategic uses** of these results include pricing optimization, product realignment, and targeted marketing.
""")

# Additional Cluster Validity Metrics
st.write("### Additional Cluster Validity Metrics")
calinski_harabasz_kmeans = calinski_harabasz_score(X_pca, data['Cluster_KMeans'])
calinski_harabasz_agg = calinski_harabasz_score(X_pca, data['Cluster_Agglomerative'])
davies_bouldin_kmeans = davies_bouldin_score(X_pca, data['Cluster_KMeans'])
davies_bouldin_agg = davies_bouldin_score(X_pca, data['Cluster_Agglomerative'])

st.write(f"Calinski-Harabasz Index (K-Means): {calinski_harabasz_kmeans:.2f}")
st.write(f"Calinski-Harabasz Index (Agglomerative): {calinski_harabasz_agg:.2f}")
st.write(f"Davies-Bouldin Index (K-Means): {davies_bouldin_kmeans:.2f}")
st.write(f"Davies-Bouldin Index (Agglomerative): {davies_bouldin_agg:.2f}")

st.write("### Clustering Results")
st.write("#### Plain Results")

# Dosya yolunu belirt
file_path = "sheets/step6_final_results.xlsx"


with open(file_path, "rb") as file:
    file_data = file.read()

# İndirme butonunu oluştur
st.download_button(
    label="Download Plain Results as Excel",
    data=file_data,
    file_name="plain_results.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.write("#### Colored Results")

# Dosya yolunu belirt
file_path = "sheets/step6_final_results_colored.xlsx"

with open(file_path, "rb") as file:
    file_data = file.read()

# İndirme butonunu oluştur
st.download_button(
    label="Download Colored Results as Excel",
    data=file_data,
    file_name="colored_results.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


